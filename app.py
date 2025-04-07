import sqlite3
from flask import Flask, request, redirect, send_file, url_for
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

app = Flask(__name__)


def init_db():
    with sqlite3.connect("an30.db") as conn:
        c = conn.cursor()
        c.execute("PRAGMA foreign_keys = ON")
        c.execute("""CREATE TABLE IF NOT EXISTS machines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS drivers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS counterparties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date DATE NOT NULL,
            machine_id INTEGER NOT NULL,
            driver_id INTEGER NOT NULL,
            start_time TEXT,
            end_time TEXT,
            hours INTEGER DEFAULT 0,
            comment TEXT,
            counterparty_id INTEGER,
            status TEXT NOT NULL CHECK(status IN ('work', 'stop', 'repair', 'holiday')),
            FOREIGN KEY(machine_id) REFERENCES machines(id) ON DELETE CASCADE,
            FOREIGN KEY(driver_id) REFERENCES drivers(id) ON DELETE CASCADE,
            FOREIGN KEY(counterparty_id) REFERENCES counterparties(id) ON DELETE SET NULL
        )""")
        conn.commit()

app.secret_key = 'supersecretkey123'
app.config['DATABASE'] = 'an30.db'
app.config['SQLITE_TIMEOUT'] = 20

COLORS = {
    'primary': "#6C7A89",
    'secondary': "#95A5A6",
    'background': "#F5F7FA",
    'accent': "#4A90E2",
    'danger': "#ff4444",
    'status': {
        'work':    "#C8E6C9",
        'stop':    "#FFCDD2",
        'repair':  "#FFF9C4",
        'holiday': "#E1BEE7"
    }
}

RECORDS_PER_PAGE = 10  # Пагинация: число записей на странице

def init_db():
    with app.app_context():
        conn = sqlite3.connect(app.config['DATABASE'], timeout=app.config['SQLITE_TIMEOUT'])
        conn.execute("PRAGMA foreign_keys = ON")
        c = conn.cursor()

        # Раскомментировать при необходимости пересоздания таблиц (удалит данные!):
        # c.execute("DROP TABLE IF EXISTS records")
        # c.execute("DROP TABLE IF EXISTS machines")
        # c.execute("DROP TABLE IF EXISTS drivers")
        # c.execute("DROP TABLE IF EXISTS counterparties")

        c.execute('''
            CREATE TABLE IF NOT EXISTS machines (
                id INTEGER PRIMARY KEY,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS drivers (
                id INTEGER PRIMARY KEY,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS counterparties (
                id INTEGER PRIMARY KEY,
                name TEXT UNIQUE NOT NULL
            )
        ''')
        c.execute('''
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY,
                date DATE NOT NULL,
                machine_id INTEGER,
                driver_id INTEGER,
                start_time TEXT,
                end_time TEXT,
                hours INTEGER DEFAULT 0,
                comment TEXT,
                counterparty_id INTEGER,
                status TEXT NOT NULL CHECK(status IN ('work', 'stop', 'repair', 'holiday')),
                FOREIGN KEY(machine_id) REFERENCES machines(id) ON DELETE SET NULL,
                FOREIGN KEY(driver_id) REFERENCES drivers(id) ON DELETE SET NULL,
                FOREIGN KEY(counterparty_id) REFERENCES counterparties(id) ON DELETE SET NULL
            )
        ''')
        conn.commit()
        conn.close()

def get_db():
    conn = sqlite3.connect(app.config['DATABASE'], timeout=app.config['SQLITE_TIMEOUT'])
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def get_next_free_id(conn, table_name: str) -> int:
    rows = conn.execute(f"SELECT id FROM {table_name} ORDER BY id").fetchall()
    used = {r[0] for r in rows}
    candidate = 1
    while candidate in used:
        candidate += 1
    return candidate

def render_base(content):
    """Главный шаблон со стилями и отступами."""
    return f'''<!DOCTYPE html>
<html>
<head>
    <title>АН-30 Учёт</title>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        * {{
            box-sizing: border-box; margin: 0; padding: 0;
        }}
        body {{
            font-family: 'Segoe UI', sans-serif;
            background: {COLORS['background']};
            color: {COLORS['primary']};
        }}
        .header {{
            background: {COLORS['primary']};
            padding: 1rem;
            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        }}
        .nav {{
            max-width: 1200px;
            margin: 0 auto;
            display: flex;
            gap: 1rem;
        }}
        .nav a {{
            color: white;
            text-decoration: none;
            padding: 0.5rem 1rem;
            border-radius: 4px;
            transition: 0.3s;
            font-weight: bold;
        }}
        .nav a:hover {{
            background: {COLORS['secondary']};
        }}
        .container {{
            max-width: 1200px;
            margin: 2rem auto;
            padding: 0 1rem;
        }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 1.5rem;
            box-shadow: 0 2px 5px rgba(0,0,0,0.05);
            margin-bottom: 1rem;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }}
        th, td {{
            padding: 1rem;
            text-align: left;
            border-bottom: 1px solid #eee;
        }}
        th {{
            background: {COLORS['primary']};
            color: white;
            cursor: pointer;
        }}
        .status {{
            display: inline-block;
            padding: 0.25rem 0.75rem;
            border-radius: 1rem;
            font-size: 0.9em;
        }}
        .btn {{
            background: {COLORS['accent']};
            color: white;
            padding: 0.5rem 1rem;
            border: none;
            border-radius: 4px;
            cursor: pointer;
            transition: 0.3s;
            font-weight: bold;
            margin: 0 0.25rem 0.25rem 0;
        }}
        .btn-danger {{
            background: {COLORS['danger']} !important;
        }}
        .btn:hover {{
            opacity: 0.9;
        }}
        .back-btn {{
            background: {COLORS['secondary']};
            margin: 1rem 0; 
        }}
        form {{
            display: flex;
            gap: 1rem;
            flex-wrap: wrap;
        }}
        input, select {{
            padding: 0.5rem;
            border: 1px solid #ddd;
            border-radius: 4px;
            min-width: 150px;
        }}
        .action-buttons {{
            display: inline-flex;
            gap: 0.5rem;
        }}
        .calendar-grid {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 0.5rem;
            margin-top: 1rem;
        }}
        .calendar-day {{
            background: white;
            padding: 1rem;
            border-radius: 8px;
            min-height: 120px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .filters {{
            margin-bottom: 1rem;
        }}
        .pagination {{
            margin-top: 1rem;
            display: flex;
            gap: 0.5rem;
        }}
        .pagination span {{
            padding: 0.5rem 1rem;
        }}

        /* Специально для отступов в заголовках/кнопках */
        .calendar-header {{
            display: flex;
            align-items: center;
            gap: 1rem;
            flex-wrap: wrap;
            margin-top: 1rem;
        }}
        .calendar-header h1 {{
            margin: 0;
            font-size: 1.5rem;
        }}
        .calendar-nav-btns {{
            display: inline-flex;
            gap: 1rem;
        }}
    </style>
</head>
<body>
    <header class="header">
        <nav class="nav">
            <a href="/">Главная</a>
            <a href="/admin">Админка</a>
            <a href="/export">&#128202; Отчёт (все)</a>
        </nav>
    </header>
    <div class="container">
        {content}
    </div>
    <script>
        function confirmDelete(msg) {{
            return confirm(msg || 'Вы уверены что хотите удалить запись?');
        }}
        function sortBy(sortField) {{
            const url = new URL(window.location.href);
            let currentSort = url.searchParams.get('sort');
            if (currentSort === sortField + '_asc') {{
                url.searchParams.set('sort', sortField + '_desc');
            }} else {{
                url.searchParams.set('sort', sortField + '_asc');
            }}
            window.location.href = url.toString();
        }}
    </script>
</body>
</html>'''

# --------------------- ВСТАВКА / УТИЛИТЫ ---------------------

def insert_machine(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "machines")
        conn.execute("INSERT INTO machines (id,name) VALUES (?,?)", (new_id,name))
        conn.commit()
    finally:
        conn.close()

def insert_driver(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "drivers")
        conn.execute("INSERT INTO drivers (id,name) VALUES (?,?)", (new_id,name))
        conn.commit()
    finally:
        conn.close()

def insert_counterparty(name: str):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "counterparties")
        conn.execute("INSERT INTO counterparties (id,name) VALUES (?,?)", (new_id,name))
        conn.commit()
    finally:
        conn.close()

def insert_record(date_str, machine_id, driver_id, status, start_time, end_time, hours, comment, counterparty_id):
    conn = get_db()
    try:
        new_id = get_next_free_id(conn, "records")
        conn.execute('''
            INSERT INTO records
            (id,date,machine_id,driver_id,status,start_time,end_time,hours,comment,counterparty_id)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        ''',(new_id,date_str,machine_id,driver_id,status,start_time,end_time,hours,comment,counterparty_id))
        conn.commit()
    finally:
        conn.close()

# --------------------- ГЛАВНАЯ ---------------------

@app.route('/')
def index():
    conn = get_db()
    try:
        machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    finally:
        conn.close()
    rows = ""
    for m in machines:
        rows += f"""
        <tr>
            <td>{m[1]}</td>
            <td><a class="btn" href="/calendar/{m[0]}">&#128197; Календарь</a></td>
        </tr>
        """
    return render_base(f'''
        <div class="card">
            <h1>Учёт работы спецтехники</h1>
            <table>
                <tr><th>Техника</th><th>Действия</th></tr>
                {rows}
            </table>
        </div>
    ''')

# --------------------- КАЛЕНДАРЬ ---------------------

@app.route('/calendar/<int:machine_id>')
def calendar(machine_id):
    year = request.args.get('year', type=int, default=datetime.now().year)
    month= request.args.get('month',type=int, default=datetime.now().month)
    if month<1: month=1
    if month>12: month=12
    if year<2020: year=2020
    if year>2030: year=2030

    conn = get_db()
    try:
        machine = conn.execute("SELECT * FROM machines WHERE id=?", (machine_id,)).fetchone()
        if not machine:
            return render_base("<h2>Техника не найдена</h2>"),404

        first_day = datetime(year,month,1)
        last_day  = (first_day.replace(day=28)+timedelta(days=4)).replace(day=1)-timedelta(days=1)
        dates = [first_day+timedelta(days=i) for i in range((last_day-first_day).days+1)]

        recs_dict = {}
        for d in dates:
            recs = conn.execute('''
                SELECT IFNULL(d.name,"Водитель удалён"), r.status, r.start_time, r.end_time, IFNULL(c.name,"")
                  FROM records r
             LEFT JOIN drivers d ON r.driver_id=d.id
             LEFT JOIN counterparties c ON r.counterparty_id=c.id
                 WHERE r.machine_id=? AND r.date=?
            ''',(machine_id,d.date())).fetchall()
            recs_dict[d.date()] = recs
    finally:
        conn.close()

    prev_month = month-1
    prev_year  = year
    if prev_month<1:
        prev_month=12
        prev_year-=1

    next_month = month+1
    next_year  = year
    if next_month>12:
        next_month=1
        next_year+=1

    # Блок кнопок и заголовка оформляем с отступами
    calendar_nav = f'''
    <div class="calendar-header">
        <div style="flex:1;">
            <h1 style="margin-bottom:0;">{machine[1]}</h1>
            <div style="font-size:1rem;color:{COLORS['secondary']};">
                {first_day.strftime("%B %Y")}
            </div>
        </div>
        <div class="calendar-nav-btns">
            <a class="btn" href="/calendar/{machine_id}?year={prev_year}&month={prev_month}">
                ← Пред. месяц
            </a>
            <a class="btn" href="/calendar/{machine_id}?year={next_year}&month={next_month}">
                След. месяц →
            </a>
        </div>
    </div>
    '''

    cal_html = '<div class="calendar-grid">'
    for d in dates:
        day_recs = recs_dict.get(d.date(), [])
        inside = ""
        for r in day_recs:
            driver_ = r[0]
            status_ = r[1]
            st = r[2] or ""
            en = r[3] or ""
            cparty_ = r[4]
            color_ = COLORS['status'].get(status_,"#fff")
            inside += f'''
            <div class="status" style="background:{color_};margin-bottom:0.5rem;">
                {driver_} - {status_.capitalize()}<br>
                {f"{st} - {en}" if st and en else ""}
                <br>{cparty_}
            </div>
            '''
        cal_html += f'''
        <div class="calendar-day">
            <div style="font-weight:bold;margin-bottom:0.5rem;font-size:1.1rem;">
                {d.strftime("%d.%m")}
            </div>
            {inside}
        </div>
        '''
    cal_html+='</div>'

    return render_base(f'''
        <a href="/" class="btn back-btn">← Назад</a>
        <div class="card" style="margin-top:1rem;">
            {calendar_nav}
            {cal_html}
        </div>
    ''')

# --------------------- АДМИНКА ---------------------

@app.route('/admin')
def admin():
    return render_base('''
        <div class="card">
            <h1>Административная панель</h1>
            <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:1rem;">
                <a class="btn" href="/admin/machines">&#128668; Техника</a>
                <a class="btn" href="/admin/drivers">&#128100; Водители</a>
                <a class="btn" href="/admin/counterparties">&#127970; Контрагенты</a>
                <a class="btn" href="/admin/records">&#128197; Записи</a>
            </div>
        </div>
    ''')

# --------------------- МАШИНЫ ---------------------

@app.route('/admin/machines', methods=['GET','POST'])
def admin_machines():
    if request.method=='POST':
        insert_machine(request.form['name'])
        return redirect('/admin/machines')

    conn = get_db()
    try:
        machines = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    finally:
        conn.close()

    rows = ""
    for m in machines:
        rows += f'''
        <tr>
            <td>{m[0]}</td>
            <td>{m[1]}</td>
            <td class="action-buttons">
                <a href="/edit/machine/{m[0]}" class="btn">Редактировать</a>
                <form method="POST" action="/delete/machine/{m[0]}">
                    <button type="submit" class="btn btn-danger"
                            onclick="return confirmDelete('Удалить машину {m[1]}?')">
                        Удалить
                    </button>
                </form>
            </td>
        </tr>
        '''
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление техникой</h1>
            <form method="POST" style="margin-bottom:1rem;">
                <input type="text" name="name" placeholder="Название техники" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Название</th><th>Действия</th></tr>
                {rows}
            </table>
        </div>
    ''')

@app.route('/edit/machine/<int:id>', methods=['GET','POST'])
def edit_machine(id):
    conn = get_db()
    if request.method=='POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE machines SET name=? WHERE id=?", (new_name,id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/machines')
    else:
        machine = conn.execute("SELECT * FROM machines WHERE id=?", (id,)).fetchone()
        conn.close()
        if not machine:
            return render_base("<h2>Машина не найдена</h2>"),404
        return render_base(f'''
            <a href="/admin/machines" class="btn back-btn">← Назад</a>
            <div class="card">
                <h1>Редактировать технику</h1>
                <form method="POST">
                    <input type="text" name="name" value="{machine[1]}" required>
                    <button type="submit" class="btn">Сохранить</button>
                </form>
            </div>
        ''')

@app.route('/delete/machine/<int:id>', methods=['POST'])
def delete_machine(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM machines WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления",500
    finally:
        conn.close()
    return redirect('/admin/machines')

# --------------------- ВОДИТЕЛИ ---------------------

@app.route('/admin/drivers', methods=['GET','POST'])
def admin_drivers():
    if request.method=='POST':
        insert_driver(request.form['name'])
        return redirect('/admin/drivers')

    conn = get_db()
    try:
        drivers = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
    finally:
        conn.close()

    rows=""
    for d in drivers:
        rows+=f'''
        <tr>
            <td>{d[0]}</td>
            <td>{d[1]}</td>
            <td class="action-buttons">
                <a href="/edit/driver/{d[0]}" class="btn">Редактировать</a>
                <form method="POST" action="/delete/driver/{d[0]}">
                    <button type="submit" class="btn btn-danger" 
                            onclick="return confirmDelete('Удалить водителя {d[1]}?')">
                        Удалить
                    </button>
                </form>
            </td>
        </tr>
        '''
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление водителями</h1>
            <form method="POST" style="margin-bottom:1rem;">
                <input type="text" name="name" placeholder="ФИО водителя" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Имя</th><th>Действия</th></tr>
                {rows}
            </table>
        </div>
    ''')

@app.route('/edit/driver/<int:id>', methods=['GET','POST'])
def edit_driver(id):
    conn = get_db()
    if request.method=='POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE drivers SET name=? WHERE id=?", (new_name,id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/drivers')
    else:
        driver = conn.execute("SELECT * FROM drivers WHERE id=?", (id,)).fetchone()
        conn.close()
        if not driver:
            return render_base("<h2>Водитель не найден</h2>"),404
        return render_base(f'''
            <a href="/admin/drivers" class="btn back-btn">← Назад</a>
            <div class="card">
                <h1>Редактировать водителя</h1>
                <form method="POST">
                    <input type="text" name="name" value="{driver[1]}" required>
                    <button type="submit" class="btn">Сохранить</button>
                </form>
            </div>
        ''')

@app.route('/delete/driver/<int:id>', methods=['POST'])
def delete_driver(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM drivers WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления",500
    finally:
        conn.close()
    return redirect('/admin/drivers')

# --------------------- КОНТРАГЕНТЫ ---------------------

@app.route('/admin/counterparties', methods=['GET','POST'])
def admin_counterparties():
    if request.method=='POST':
        insert_counterparty(request.form['name'])
        return redirect('/admin/counterparties')

    conn = get_db()
    try:
        cparties = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
    finally:
        conn.close()

    rows=""
    for cp in cparties:
        rows+=f'''
        <tr>
            <td>{cp[0]}</td>
            <td>{cp[1]}</td>
            <td class="action-buttons">
                <a href="/edit/counterparty/{cp[0]}" class="btn">Редактировать</a>
                <form method="POST" action="/delete/counterparty/{cp[0]}">
                    <button type="submit" class="btn btn-danger"
                            onclick="return confirmDelete('Удалить контрагента {cp[1]}?')">
                        Удалить
                    </button>
                </form>
            </td>
        </tr>
        '''
    return render_base(f'''
        <a href="/admin" class="btn back-btn">← Назад</a>
        <div class="card">
            <h1>Управление контрагентами</h1>
            <form method="POST" style="margin-bottom:1rem;">
                <input type="text" name="name" placeholder="Название контрагента" required>
                <button type="submit" class="btn">Добавить</button>
            </form>
            <table>
                <tr><th>ID</th><th>Название</th><th>Действия</th></tr>
                {rows}
            </table>
        </div>
    ''')

@app.route('/edit/counterparty/<int:id>', methods=['GET','POST'])
def edit_counterparty(id):
    conn = get_db()
    if request.method=='POST':
        new_name = request.form['name']
        try:
            conn.execute("UPDATE counterparties SET name=? WHERE id=?", (new_name,id))
            conn.commit()
        except:
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/counterparties')
    else:
        cp = conn.execute("SELECT * FROM counterparties WHERE id=?", (id,)).fetchone()
        conn.close()
        if not cp:
            return render_base("<h2>Контрагент не найден</h2>"),404
        return render_base(f'''
            <a href="/admin/counterparties" class="btn back-btn">← Назад</a>
            <div class="card">
                <h1>Редактировать контрагента</h1>
                <form method="POST">
                    <input type="text" name="name" value="{cp[1]}" required>
                    <button type="submit" class="btn">Сохранить</button>
                </form>
            </div>
        ''')

@app.route('/delete/counterparty/<int:id>', methods=['POST'])
def delete_counterparty(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM counterparties WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления",500
    finally:
        conn.close()
    return redirect('/admin/counterparties')

# --------------------- ЗАПИСИ (СПРАВА - ФИЛЬТРЫ), ПРИ ЭТОМ ОФОРМЛЕНИЕ ОПРЯТНОЕ ---------------------

@app.route('/admin/records', methods=['GET','POST'])
def admin_records():
    if request.method=='POST':
        # Добавить запись
        date_str = request.form['date']
        machine_id=int(request.form['machine_id'])
        driver_id =int(request.form['driver_id'])
        status= request.form['status']
        start_t= request.form.get('start_time','')
        end_t  = request.form.get('end_time','')
        comm   = request.form.get('comment','')
        c_id   = request.form.get('counterparty_id')
        cpar_id= int(c_id) if c_id else None

        hours=0
        if start_t and end_t:
            try:
                st=datetime.strptime(start_t,'%H:%M')
                en=datetime.strptime(end_t,'%H:%M')
                if en<st: en+=timedelta(days=1)
                hours=(en-st).seconds//3600
            except:
                pass

        insert_record(date_str,machine_id,driver_id,status,start_t or None,end_t or None,hours,comm,cpar_id)
        return redirect('/admin/records')

    # GET
    # Фильтры
    date_from=request.args.get('date_from','')
    date_to=  request.args.get('date_to','')
    mach_f=  request.args.get('mach', type=int)
    driv_f=  request.args.get('driv', type=int)
    cpar_f=  request.args.get('cpar', type=int)
    stat_f=  request.args.get('status','')
    comm_sub=request.args.get('comment_sub','').strip()
    sort_key=request.args.get('sort','date_desc')
    page=    request.args.get('page', type=int, default=1)
    if page<1: page=1

    where=[]
    pr=[]
    if date_from:
        where.append("r.date>=?")
        pr.append(date_from)
    if date_to:
        where.append("r.date<=?")
        pr.append(date_to)
    if mach_f:
        where.append("r.machine_id=?")
        pr.append(mach_f)
    if driv_f:
        where.append("r.driver_id=?")
        pr.append(driv_f)
    if cpar_f:
        where.append("r.counterparty_id=?")
        pr.append(cpar_f)
    if stat_f in ("work","stop","repair","holiday"):
        where.append("r.status=?")
        pr.append(stat_f)
    if comm_sub:
        where.append("r.comment LIKE ?")
        pr.append(f"%{comm_sub}%")

    where_sql=""
    if where:
        where_sql="WHERE "+ " AND ".join(where)

    # Сортировка
    if sort_key=="date_asc":
        order_sql="ORDER BY r.date ASC, r.id ASC"
    elif sort_key=="date_desc":
        order_sql="ORDER BY r.date DESC, r.id DESC"
    elif sort_key=="hours_asc":
        order_sql="ORDER BY r.hours ASC, r.date ASC"
    elif sort_key=="hours_desc":
        order_sql="ORDER BY r.hours DESC, r.date DESC"
    elif sort_key=="machine_asc":
        order_sql="ORDER BY m.name ASC, r.date DESC"
    elif sort_key=="driver_asc":
        order_sql="ORDER BY d.name ASC, r.date DESC"
    else:
        order_sql="ORDER BY r.date DESC, r.id DESC"

    conn = get_db()
    # Для пагинации
    count_sql=f'''
        SELECT COUNT(*)
          FROM records r
     LEFT JOIN machines m ON r.machine_id=m.id
     LEFT JOIN drivers d ON r.driver_id=d.id
     LEFT JOIN counterparties c ON r.counterparty_id=c.id
        {where_sql}
    '''
    total_count=conn.execute(count_sql, pr).fetchone()[0]
    total_pages=(total_count+RECORDS_PER_PAGE-1)//RECORDS_PER_PAGE
    offset=(page-1)*RECORDS_PER_PAGE

    query=f'''
        SELECT r.id,
               r.date,
               IFNULL(m.name,"Техника удал/не выбрана"),
               IFNULL(d.name,"Водитель удал/не выбран"),
               r.start_time,
               r.end_time,
               r.hours,
               IFNULL(r.comment,"-"),
               IFNULL(c.name,"Контрагента нет"),
               r.status
          FROM records r
     LEFT JOIN machines m ON r.machine_id=m.id
     LEFT JOIN drivers d ON r.driver_id=d.id
     LEFT JOIN counterparties c ON r.counterparty_id=c.id
        {where_sql}
        {order_sql}
        LIMIT ? OFFSET ?
    '''
    recs=conn.execute(query, pr+[RECORDS_PER_PAGE, offset]).fetchall()

    machines   = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
    drivers    = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
    cparties   = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
    conn.close()

    # Список options
    mach_opts="".join(f'<option value="{m[0]}" {"selected" if mach_f==m[0] else ""}>{m[1]}</option>' for m in machines)
    driv_opts="".join(f'<option value="{d[0]}" {"selected" if driv_f==d[0] else ""}>{d[1]}</option>' for d in drivers)
    cpar_opts="".join(f'<option value="{c[0]}" {"selected" if cpar_f==c[0] else ""}>{c[1]}</option>' for c in cparties)

    # Форма фильтров - справа
    def sel(a,b): return "selected" if a==b else ""
    filters_html=f'''
    <div class="card" style="margin-bottom:1rem;">
        <h2>Фильтры</h2>
        <form method="GET" style="flex-direction:column;gap:0.5rem;">
            <label>Дата с:</label>
            <input type="date" name="date_from" value="{date_from}">
            <label>Дата по:</label>
            <input type="date" name="date_to" value="{date_to}">
            <label>Техника:</label>
            <select name="mach">
                <option value="">[Все]</option>
                {mach_opts}
            </select>
            <label>Водитель:</label>
            <select name="driv">
                <option value="">[Все]</option>
                {driv_opts}
            </select>
            <label>Контрагент:</label>
            <select name="cpar">
                <option value="">[Все]</option>
                {cpar_opts}
            </select>
            <label>Статус:</label>
            <select name="status">
                <option value="">[Все]</option>
                <option value="work" {sel("work",stat_f)}>Работа</option>
                <option value="stop" {sel("stop",stat_f)}>Простой</option>
                <option value="repair" {sel("repair",stat_f)}>Ремонт</option>
                <option value="holiday" {sel("holiday",stat_f)}>Выходной</option>
            </select>
            <label>Комментарий (поиск):</label>
            <input type="text" name="comment_sub" value="{comm_sub}">
            <button type="submit" class="btn" style="margin-top:1rem;">Применить</button>
            <a class="btn" href="{url_for('export_excel')}?export=filtered
                &date_from={date_from}&date_to={date_to}
                &mach={mach_f or ''}&driv={driv_f or ''}&cpar={cpar_f or ''}
                &status={stat_f}&comment_sub={comm_sub}&sort={sort_key}">
                Экспорт
            </a>
        </form>
    </div>
    '''

    # Форма добавления слева
    create_form_html=f'''
    <div class="card">
        <h2>Добавить новую запись</h2>
        <form method="POST">
            <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:1rem;">
                <input type="date" name="date" required>
                <select name="machine_id" required>
                    <option value="">Выберите технику</option>
                    {"".join(f'<option value="{m[0]}">{m[1]}</option>' for m in machines)}
                </select>
                <select name="driver_id" required>
                    <option value="">Выберите водителя</option>
                    {"".join(f'<option value="{d[0]}">{d[1]}</option>' for d in drivers)}
                </select>
                <select name="status" required>
                    <option value="work">Работа</option>
                    <option value="stop">Простой</option>
                    <option value="repair">Ремонт</option>
                    <option value="holiday">Выходной</option>
                </select>
                <input type="time" name="start_time" placeholder="Начало">
                <input type="time" name="end_time"   placeholder="Конец">
                <select name="counterparty_id">
                    <option value="">Контрагент (не обязательно)</option>
                    {"".join(f'<option value="{c[0]}">{c[1]}</option>' for c in cparties)}
                </select>
                <input type="text" name="comment" placeholder="Комментарий" style="grid-column:span 2;">
            </div>
            <button type="submit" class="btn" style="width:100%;margin-top:1rem;">Добавить запись</button>
        </form>
    </div>
    '''

    # Список записей
    rows_html=""
    for r in recs:
        rec_id=r[0]
        date_db=r[1]
        mach_nm=r[2]
        driv_nm=r[3]
        st=r[4] or ""
        en=r[5] or ""
        hrs=r[6] or 0
        comm=r[7]
        cpar=r[8]
        stat_=r[9]
        try:
            date_fmt=datetime.strptime(date_db,'%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_fmt=date_db
        time_str=f"{st} - {en}" if (st and en) else "-"
        color=COLORS['status'].get(stat_,"#fff")
        rows_html+=f'''
        <tr>
            <td>{date_fmt}</td>
            <td>{mach_nm}</td>
            <td>{driv_nm}</td>
            <td>{time_str}</td>
            <td>{hrs}</td>
            <td>{cpar}</td>
            <td>{comm}</td>
            <td>
                <div class="status" style="background:{color};">
                    {stat_.capitalize()}
                </div>
            </td>
            <td class="action-buttons">
                <a href="/edit/record/{rec_id}" class="btn">Редактировать</a>
                <form method="POST" action="/delete/record/{rec_id}">
                    <button type="submit" class="btn btn-danger" 
                            onclick="return confirmDelete('Удалить запись?')">
                        Удалить
                    </button>
                </form>
            </td>
        </tr>
        '''

    pagination_html=""
    if total_pages>1:
        pagination_html+='<div class="pagination">'
        if page>1:
            prevp=page-1
            qstr=request.query_string.decode("utf-8")
            pagination_html+=f'<a class="btn" href="?{qstr.replace(f"page={page}",f"page={prevp}")}">←</a>'
        else:
            pagination_html+='<span>←</span>'
        pagination_html+=f'<span>Стр. {page}/{total_pages}</span>'
        if page<total_pages:
            nextp=page+1
            qstr=request.query_string.decode("utf-8")
            if "page=" in qstr:
                pagination_html+=f'<a class="btn" href="?{qstr.replace(f"page={page}",f"page={nextp}")}">→</a>'
            else:
                pagination_html+=f'<a class="btn" href="?{qstr}&page={nextp}">→</a>'
        else:
            pagination_html+='<span>→</span>'
        pagination_html+='</div>'

    table_html=f'''
    <div class="card" style="margin-top:1rem;">
        <h2>Список записей</h2>
        <table style="margin-top:1rem;">
            <tr>
                <th onclick="sortBy('date')">Дата</th>
                <th onclick="sortBy('machine')">Техника</th>
                <th onclick="sortBy('driver')">Водитель</th>
                <th>Время</th>
                <th onclick="sortBy('hours')">Часы</th>
                <th>Контрагент</th>
                <th>Комментарий</th>
                <th onclick="sortBy('status')">Статус</th>
                <th>Действия</th>
            </tr>
            {rows_html}
        </table>
        {pagination_html}
    </div>
    '''

    # Размещаем всё в flex: слева добавление + таблица, справа фильтры
    content=f'''
    <a href="/admin" class="btn back-btn">← Назад</a>
    <div style="display:flex;align-items:flex-start;gap:1rem;flex-wrap:wrap;">
        <div style="flex:1;min-width:400px;">
            {create_form_html}
            {table_html}
        </div>
        <div style="width:300px;flex-shrink:0;">
            {filters_html}
        </div>
    </div>
    '''
    return render_base(content)

# --------------------- РЕДАКТИРОВАНИЕ ЗАПИСИ ---------------------

@app.route('/edit/record/<int:id>', methods=['GET','POST'])
def edit_record(id):
    conn = get_db()
    if request.method=='POST':
        try:
            date_str=request.form['date']
            machine_id=int(request.form['machine_id'])
            driver_id =int(request.form['driver_id'])
            status=   request.form['status']
            start_t=  request.form.get('start_time','')
            end_t=    request.form.get('end_time','')
            comm=     request.form.get('comment','')
            c_id=     request.form.get('counterparty_id')
            cpar_id=  int(c_id) if c_id else None

            hours=0
            if start_t and end_t:
                try:
                    st=datetime.strptime(start_t,'%H:%M')
                    en=datetime.strptime(end_t,'%H:%M')
                    if en<st: en+=timedelta(days=1)
                    hours=(en-st).seconds//3600
                except:
                    pass

            conn.execute('''
                UPDATE records
                   SET date=?,
                       machine_id=?,
                       driver_id=?,
                       status=?,
                       start_time=?,
                       end_time=?,
                       hours=?,
                       comment=?,
                       counterparty_id=?
                 WHERE id=?
            ''',(date_str,machine_id,driver_id,status,start_t or None,end_t or None,hours,comm,cpar_id,id))
            conn.commit()
        except Exception as e:
            print(f"Ошибка редактирования: {e}")
            conn.rollback()
        finally:
            conn.close()
        return redirect('/admin/records')
    else:
        record = conn.execute('''
            SELECT date,machine_id,driver_id,status,start_time,end_time,hours,comment,counterparty_id
              FROM records
             WHERE id=?
        ''',(id,)).fetchone()

        if not record:
            conn.close()
            return render_base("<h2>Запись не найдена</h2>"),404

        machines  = conn.execute("SELECT * FROM machines ORDER BY id").fetchall()
        drivers   = conn.execute("SELECT * FROM drivers ORDER BY id").fetchall()
        cparties  = conn.execute("SELECT * FROM counterparties ORDER BY id").fetchall()
        conn.close()

        date_val=record[0]
        mach_val=record[1]
        driv_val=record[2]
        stat_val=record[3]
        st_val=record[4] or ""
        en_val=record[5] or ""
        comm_val=record[7] or ""
        cpar_val=record[8]

        def sel(a,b): return "selected" if a==b else ""
        mach_opts="".join(f'<option value="{m[0]}" {sel(m[0],mach_val)}>{m[1]}</option>' for m in machines)
        driv_opts="".join(f'<option value="{d[0]}" {sel(d[0],driv_val)}>{d[1]}</option>' for d in drivers)

        status_opts=""
        for s_val, s_lbl in [('work','Работа'),('stop','Простой'),('repair','Ремонт'),('holiday','Выходной')]:
            status_opts+=f'<option value="{s_val}" {sel(s_val,stat_val)}>{s_lbl}</option>'

        cparty_opts='<option value="">Контрагент (не обязательно)</option>'
        for cp in cparties:
            cparty_opts+=f'<option value="{cp[0]}" {sel(cp[0],cpar_val)}>{cp[1]}</option>'

        return render_base(f'''
            <a href="/admin/records" class="btn back-btn">← Назад</a>
            <div class="card">
                <h2>Редактировать запись</h2>
                <form method="POST">
                    <div style="display:grid;grid-template-columns:repeat(2,1fr);gap:1rem;">
                        <input type="date" name="date" value="{date_val}" required>
                        <select name="machine_id" required>
                            {mach_opts}
                        </select>
                        <select name="driver_id" required>
                            {driv_opts}
                        </select>
                        <select name="status" required>
                            {status_opts}
                        </select>
                        <input type="time" name="start_time" value="{st_val}">
                        <input type="time" name="end_time"   value="{en_val}">
                        <select name="counterparty_id">
                            {cparty_opts}
                        </select>
                        <input type="text" name="comment" value="{comm_val}" style="grid-column:span 2;">
                    </div>
                    <button type="submit" class="btn" style="margin-top:1rem;">
                        Сохранить
                    </button>
                </form>
            </div>
        ''')

@app.route('/delete/record/<int:id>', methods=['POST'])
def delete_record(id):
    conn = get_db()
    try:
        conn.execute("DELETE FROM records WHERE id=?", (id,))
        conn.commit()
    except:
        conn.rollback()
        return "Ошибка удаления записи", 500
    finally:
        conn.close()
    return redirect('/admin/records')

# --------------------- ВЫГРУЗКА В EXCEL ---------------------

@app.route('/export')
def export_excel():
    export_mode = request.args.get('export')

    conn = get_db()
    try:
        if export_mode=='filtered':
            # Те же фильтры, что и в /admin/records
            date_from=request.args.get('date_from','')
            date_to=  request.args.get('date_to','')
            mach_f=  request.args.get('mach',  type=int)
            driv_f=  request.args.get('driv',  type=int)
            cpar_f=  request.args.get('cpar',  type=int)
            stat_f=  request.args.get('status','')
            comm_sub=request.args.get('comment_sub','').strip()
            sort_key=request.args.get('sort','date_desc')

            wh=[]
            pr=[]
            if date_from:
                wh.append("r.date>=?")
                pr.append(date_from)
            if date_to:
                wh.append("r.date<=?")
                pr.append(date_to)
            if mach_f:
                wh.append("r.machine_id=?")
                pr.append(mach_f)
            if driv_f:
                wh.append("r.driver_id=?")
                pr.append(driv_f)
            if cpar_f:
                wh.append("r.counterparty_id=?")
                pr.append(cpar_f)
            if stat_f in ("work","stop","repair","holiday"):
                wh.append("r.status=?")
                pr.append(stat_f)
            if comm_sub:
                wh.append("r.comment LIKE ?")
                pr.append(f"%{comm_sub}%")

            where_sql=""
            if wh:
                where_sql="WHERE "+ " AND ".join(wh)

            if sort_key=="date_asc":
                order_sql="ORDER BY r.date ASC, r.id ASC"
            elif sort_key=="date_desc":
                order_sql="ORDER BY r.date DESC, r.id DESC"
            elif sort_key=="hours_asc":
                order_sql="ORDER BY r.hours ASC, r.date ASC"
            elif sort_key=="hours_desc":
                order_sql="ORDER BY r.hours DESC, r.date DESC"
            elif sort_key=="machine_asc":
                order_sql="ORDER BY m.name ASC, r.date DESC"
            elif sort_key=="driver_asc":
                order_sql="ORDER BY d.name ASC, r.date DESC"
            else:
                order_sql="ORDER BY r.date DESC, r.id DESC"

            sql = f'''
                SELECT r.date,
                       IFNULL(m.name,"Техника нет/удалена"),
                       IFNULL(d.name,"Водитель нет/удалён"),
                       r.status,
                       IFNULL(r.start_time,""),
                       IFNULL(r.end_time,""),
                       r.hours,
                       IFNULL(c.name,"Контрагента нет"),
                       IFNULL(r.comment,"-")
                  FROM records r
             LEFT JOIN machines m ON r.machine_id=m.id
             LEFT JOIN drivers d ON r.driver_id=d.id
             LEFT JOIN counterparties c ON r.counterparty_id=c.id
                {where_sql}
                {order_sql}
            '''
            rows=conn.execute(sql, pr).fetchall()
        else:
            # Все
            rows=conn.execute('''
                SELECT r.date,
                       IFNULL(m.name,"Техника нет/удалена"),
                       IFNULL(d.name,"Водитель нет/удалён"),
                       r.status,
                       IFNULL(r.start_time,""),
                       IFNULL(r.end_time,""),
                       r.hours,
                       IFNULL(c.name,"Контрагента нет"),
                       IFNULL(r.comment,"-")
                  FROM records r
             LEFT JOIN machines m ON r.machine_id=m.id
             LEFT JOIN drivers d ON r.driver_id=d.id
             LEFT JOIN counterparties c ON r.counterparty_id=c.id
              ORDER BY r.date ASC, r.id ASC
            ''').fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title="AN-30 Отчёт"

    headers=["Дата","Техника","Водитель","Статус","Начало","Конец","Часы","Контрагент","Комментарий"]
    ws.append(headers)

    header_fill=PatternFill(start_color="444444", fill_type="solid")
    header_font=Font(color="FFFFFF", bold=True)

    for col in range(1,len(headers)+1):
        cell=ws.cell(row=1,column=col)
        cell.fill=header_fill
        cell.font=header_font
        ws.column_dimensions[get_column_letter(col)].width=20

    for row in rows:
        # row => date, machine, driver, status, start, end, hours, cparty, comment
        date_db=row[0]
        try:
            date_fmt=datetime.strptime(date_db,'%Y-%m-%d').strftime('%d.%m.%Y')
        except:
            date_fmt=date_db
        machine=row[1]
        driver=row[2]
        status_=row[3]
        st_=row[4]
        en_=row[5]
        hrs_=row[6]
        cpar_=row[7]
        comm_=row[8]
        color_hex=COLORS['status'].get(status_,"#FFFFFF")[1:]

        ws.append([date_fmt,machine,driver,status_.capitalize(),st_,en_,hrs_,cpar_,comm_])
        scell=ws.cell(row=ws.max_row,column=4) # столбец "Статус"
        scell.fill=PatternFill(start_color=color_hex, fill_type="solid")

    filename="report_"+datetime.now().strftime("%Y%m%d_%H%M")+".xlsx"
    wb.save(filename)
    return send_file(
        filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__=='__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
